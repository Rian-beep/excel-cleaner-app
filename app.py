import streamlit as st
import pandas as pd
import re
from unidecode import unidecode
from ftfy import fix_text
import requests
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import json

# Try to import optional dependencies
try:
    from email_validator import validate_email, EmailNotValidError
    EMAIL_VALIDATOR_AVAILABLE = True
except ImportError:
    EMAIL_VALIDATOR_AVAILABLE = False

try:
    import phonenumbers
    from phonenumbers import NumberParseException
    PHONENUMBERS_AVAILABLE = True
except ImportError:
    PHONENUMBERS_AVAILABLE = False

# --- Page Config ---
st.set_page_config(page_title="Cleanr", layout="wide", page_icon="🧹")

# --- Global Styles ---
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;700&display=swap');

        html, body, [class*="css"] {
            font-family: 'Montserrat', sans-serif;
            background-color: #ffffff;
            color: #0a2342;
        }
        .title-text {
            text-align: center;
            font-size: 3em;
            font-weight: 700;
            margin-bottom: 0.1em;
            color: #0a2342;
        }
        .subtitle-text {
            text-align: center;
            font-size: 1.2em;
            margin-bottom: 2em;
            color: #0a2342;
        }
        .rounded-box {
            background-color: #F0F0F0;
            border-radius: 16px;
            padding: 1.5rem;
            margin-bottom: 1.5rem;
            color: #0a2342;
        }
        .feedback-container {
            background-color: #F0F0F0;
            border-radius: 16px;
            padding: 1.5rem;
            margin-top: 2rem;
            color: #0a2342;
        }
        .feedback-box textarea {
            background-color: #112340;
            color: #ffffff;
            border-radius: 12px;
        }
        .download-button button, .stButton>button, .stDownloadButton>button {
            background-color: #112340 !important;
            color: #ffffff !important;
            border-radius: 12px;
        }
        .section-header {
            font-size: 1.3em;
            font-weight: 600;
            color: #0a2342;
            margin-bottom: 0.5em;
        }
        .metric-card {
            background-color: #f8f9fa;
            padding: 1rem;
            border-radius: 8px;
            border-left: 4px solid #112340;
        }
    </style>
""", unsafe_allow_html=True)

# --- Load Known Company Names ---
company_dict = {}
company_file = "company_directory.csv"
if os.path.exists(company_file):
    known_companies_df = pd.read_csv(company_file)
    company_dict = dict(
        zip(
            known_companies_df["Raw Company"].str.strip().str.lower(),
            known_companies_df["Cleaned Company"].str.strip()
        )
    )

# --- Cleaning Rules ---
COMMON_SUFFIXES = ['ltd', 'inc', 'group', 'brands', 'company', 'companies', 'incorporation', 'corporation']

# Common disposable email domains
DISPOSABLE_EMAIL_DOMAINS = {
    '10minutemail.com', 'tempmail.com', 'guerrillamail.com', 'mailinator.com',
    'throwaway.email', 'temp-mail.org', 'getnada.com', 'mohmal.com'
}

# Common job title abbreviations
JOB_TITLE_ABBREVIATIONS = {
    'ceo': 'Chief Executive Officer',
    'cto': 'Chief Technology Officer',
    'cfo': 'Chief Financial Officer',
    'cmo': 'Chief Marketing Officer',
    'coo': 'Chief Operating Officer',
    'vp': 'Vice President',
    'svp': 'Senior Vice President',
    'evp': 'Executive Vice President',
    'dir': 'Director',
    'mgr': 'Manager',
    'sr': 'Senior',
    'jr': 'Junior',
    'eng': 'Engineer',
    'dev': 'Developer',
    'pm': 'Product Manager',
    'hr': 'Human Resources',
    'pr': 'Public Relations',
    'it': 'Information Technology'
}


def detect_email_pattern(email, first_name='', last_name=''):
    """
    Detect the pattern type of an email address.
    Returns a pattern identifier string.
    """
    if not email or '@' not in email:
        return 'unknown'
    
    email_lower = str(email).lower().strip()
    local_part = email_lower.split('@')[0]
    
    first_lower = str(first_name).lower().strip() if first_name else ''
    last_lower = str(last_name).lower().strip() if last_name else ''
    
    # Pattern: firstname.lastname
    if first_lower and last_lower:
        if f"{first_lower}.{last_lower}" == local_part:
            return 'firstname.lastname'
        if f"{first_lower}_{last_lower}" == local_part:
            return 'firstname_lastname'
        if f"{first_lower}-{last_lower}" == local_part:
            return 'firstname-lastname'
        if f"{first_lower}{last_lower}" == local_part:
            return 'firstnamelastname'
        # Pattern: firstinitial.lastname (e.g., j.smith)
        if len(first_lower) > 0 and f"{first_lower[0]}.{last_lower}" == local_part:
            return 'firstinitial.lastname'
        if f"{first_lower[0]}{last_lower}" == local_part:
            return 'firstinitiallastname'
        # Pattern: lastname.firstname
        if f"{last_lower}.{first_lower}" == local_part:
            return 'lastname.firstname'
        if f"{last_lower}_{first_lower}" == local_part:
            return 'lastname_firstname'
    
    # Pattern: just firstname or just lastname
    if first_lower and local_part == first_lower:
        return 'firstname'
    if last_lower and local_part == last_lower:
        return 'lastname'
    
    # Pattern: numbers or other characters (less common)
    if re.search(r'\d', local_part):
        return 'with_numbers'
    
    return 'other'


def analyze_company_email_patterns(df, email_col='Email', company_col='Company', 
                                     first_name_col='First Name', last_name_col='Last Name'):
    """
    Analyze email patterns for each company and return the dominant pattern.
    Only analyzes companies with 2+ contacts.
    Returns a dict: {company_name: dominant_pattern}
    """
    company_patterns = {}
    
    if company_col not in df.columns or email_col not in df.columns:
        return company_patterns
    
    # Group by company
    for company, group in df.groupby(company_col):
        # Only analyze if company has 2+ contacts
        if len(group) < 2:
            continue
        
        # Skip if company name is empty/NaN
        if pd.isna(company) or str(company).strip() == '':
            continue
        
        pattern_counts = {}
        valid_emails = 0
        
        for idx, row in group.iterrows():
            email = row.get(email_col, '')
            first_name = row.get(first_name_col, '')
            last_name = row.get(last_name_col, '')
            
            if pd.isna(email) or str(email).strip() == '':
                continue
            
            # Check if email is valid first
            is_valid, _ = validate_email_format(email)
            if not is_valid:
                continue
            
            valid_emails += 1
            pattern = detect_email_pattern(email, first_name, last_name)
            pattern_counts[pattern] = pattern_counts.get(pattern, 0) + 1
        
        # Only set pattern if we have at least 2 valid emails
        if valid_emails >= 2:
            # Find the most common pattern
            if pattern_counts:
                dominant_pattern = max(pattern_counts, key=pattern_counts.get)
                # Only use if it represents at least 50% of emails
                if pattern_counts[dominant_pattern] / valid_emails >= 0.5:
                    company_patterns[company] = {
                        'pattern': dominant_pattern,
                        'count': pattern_counts[dominant_pattern],
                        'total': valid_emails,
                        'percentage': (pattern_counts[dominant_pattern] / valid_emails) * 100
                    }
    
    return company_patterns


def check_email_pattern_match(email, first_name, last_name, company, company_patterns):
    """
    Check if an email matches the dominant pattern for its company.
    Returns (matches_pattern, pattern_type, company_pattern)
    """
    if not company or company not in company_patterns:
        return None, None, None
    
    company_info = company_patterns[company]
    expected_pattern = company_info['pattern']
    
    detected_pattern = detect_email_pattern(email, first_name, last_name)
    
    matches = (detected_pattern == expected_pattern)
    return matches, detected_pattern, expected_pattern


def validate_email_format(email):
    """Validate email format using regex and optional email-validator library."""
    if not email or pd.isna(email):
        return False, "Missing"
    
    try:
        email_str = str(email).strip()
        if not email_str:
            return False, "Missing"
        
        email_lower = email_str.lower()
        
        # Basic regex validation
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email_lower):
            return False, "Invalid Format"
        
        # Check for disposable emails
        if '@' in email_lower:
            domain = email_lower.split('@')[1]
            if domain in DISPOSABLE_EMAIL_DOMAINS:
                return False, "Disposable Email"
        
        # Use email-validator if available for stricter validation
        if EMAIL_VALIDATOR_AVAILABLE:
            try:
                # email-validator expects the email as a string, and returns a normalized version
                result = validate_email(email_str, check_deliverability=False)
                return True, "Valid"
            except EmailNotValidError as e:
                return False, "Invalid Format"
            except Exception as e:
                # If email-validator fails for any other reason, fall back to regex validation
                return True, "Valid"
        
        return True, "Valid"
    except Exception as e:
        # If anything goes wrong, return invalid
        return False, "Invalid Format"


def clean_phone_number(phone):
    """Clean and standardize phone numbers."""
    if pd.isna(phone) or not phone:
        return '', False
    
    phone_str = str(phone).strip()
    
    # Remove common non-digit characters but keep + for international
    phone_clean = re.sub(r'[^\d+]', '', phone_str)
    
    if not phone_clean:
        return '', False
    
    # Use phonenumbers library if available
    if PHONENUMBERS_AVAILABLE:
        try:
            # Try parsing as US number first
            parsed = phonenumbers.parse(phone_clean, "US")
            if phonenumbers.is_valid_number(parsed):
                formatted = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
                return formatted, True
        except NumberParseException:
            pass
        
        # Try parsing as international
        try:
            parsed = phonenumbers.parse(phone_clean, None)
            if phonenumbers.is_valid_number(parsed):
                formatted = phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
                return formatted, True
        except NumberParseException:
            pass
    
    # Fallback: basic cleaning
    if len(phone_clean) >= 10:
        # Remove leading + if no country code
        if phone_clean.startswith('+') and len(phone_clean) <= 11:
            phone_clean = phone_clean[1:]
        return phone_clean, True
    
    return phone_clean, False


def clean_job_title(title):
    """Clean and standardize job titles."""
    if pd.isna(title) or not title:
        return ''
    
    title_str = str(title).strip()
    if not title_str:
        return ''
    
    # Fix encoding
    try:
        title_str = title_str.encode('latin1').decode('utf-8')
    except Exception:
        pass
    title_str = fix_text(title_str)
    title_str = unidecode(title_str)
    
    # Convert to title case
    title_str = title_str.title()
    
    # Expand common abbreviations
    words = title_str.split()
    expanded_words = []
    for word in words:
        word_lower = word.lower().rstrip('.')
        if word_lower in JOB_TITLE_ABBREVIATIONS:
            expanded_words.append(JOB_TITLE_ABBREVIATIONS[word_lower])
        else:
            expanded_words.append(word)
    
    title_str = ' '.join(expanded_words)
    
    # Clean up extra spaces
    title_str = re.sub(r'\s+', ' ', title_str).strip()
    
    return title_str


def clean_company(name):
    """Clean company name, using known mappings if available."""
    try:
        # Handle None, NaN, or empty values
        if name is None or pd.isna(name) or name == '':
            return ''
        
        # Convert to string and check for string representations of null
        name_str = str(name).strip()
        if not name_str or name_str.lower() in ['nan', 'none', 'null', '']:
            return ''
        
        name_key = name_str.lower()

        # Check company dictionary first
        if name_key in company_dict:
            return company_dict[name_key]

        # Try to fix encoding issues
        try:
            name_str = name_str.encode('latin1').decode('utf-8')
        except (UnicodeEncodeError, UnicodeDecodeError, AttributeError):
            pass
        
        # Fix text encoding
        name_str = fix_text(name_str)
        name_str = unidecode(name_str)

        # Remove common suffixes
        name_str = re.sub(r'\b(?:' + '|'.join(COMMON_SUFFIXES) + r')\b', '', name_str, flags=re.IGNORECASE)
        # Remove special characters except spaces, hyphens, and alphanumeric
        name_str = re.sub(r'[^A-Za-z0-9\s\-]', '', name_str)
        # Clean up multiple spaces
        name_str = re.sub(r'\s{2,}', ' ', name_str).strip()

        if not name_str:
            return ''

        # Format based on length
        if len(name_str) <= 4:
            name_str = name_str.upper()
        else:
            name_str = name_str.title()

        return name_str
    except Exception as e:
        # Return original as string if cleaning fails
        try:
            return str(name).strip() if name else ''
        except:
            return ''


def clean_first_name(name):
    """Clean first name and take the first token."""
    if pd.isna(name):
        return ''
    try:
        name = name.encode('latin1').decode('utf-8')
    except Exception:
        pass
    name = fix_text(name)
    name = unidecode(str(name)).strip()
    if not name:
        return ''
    parts = name.split()
    first = parts[0]
    return first.title()


def clean_last_name(name):
    """
    Clean last name without changing its structure.
    Keep full string, fix encoding, spacing and casing.
    """
    if pd.isna(name):
        return ''
    try:
        name = name.encode('latin1').decode('utf-8')
    except Exception:
        pass
    name = fix_text(name)
    name = unidecode(str(name)).strip()
    if not name:
        return ''

    parts = name.split()
    processed_parts = []
    for part in parts:
        if part.lower().startswith("mc") and len(part) > 2:
            part = "Mc" + part[2:].title()
        else:
            part = part.title()
        processed_parts.append(part)

    return " ".join(processed_parts)


def infer_last_from_email(first_name, email):
    """
    Infer last name from email ONLY if last name is missing.
    Handles:
      - firstname.lastname@...
      - firstname_lastname@...
      - firstname-lastname@...
      - fLastname@...
    """
    if not email or pd.isna(email) or not first_name:
        return ''

    user = email.split('@')[0].lower()
    first_lower = first_name.lower()

    # firstname[._-]lastname
    m = re.match(rf"{re.escape(first_lower)}[._-]?([a-z]+)", user)
    if m:
        return m.group(1).title()

    # fLastname pattern (e.g. jsmith)
    if user.startswith(first_lower[0]):
        guess = user[len(first_lower[0]):]
        if guess:
            return guess.title()

    return ''


def find_duplicates(df, email_col='Email', name_cols=['First Name', 'Last Name']):
    """Find duplicate contacts based on email or name combination."""
    duplicates = pd.DataFrame()
    duplicate_indices = set()
    
    if email_col in df.columns:
        # Find duplicates by email
        email_dupes = df[df.duplicated(subset=[email_col], keep=False)]
        if not email_dupes.empty:
            duplicate_indices.update(email_dupes.index)
    
    # Find duplicates by name combination
    available_name_cols = [col for col in name_cols if col in df.columns]
    if len(available_name_cols) >= 2:
        name_dupes = df[df.duplicated(subset=available_name_cols, keep=False)]
        if not name_dupes.empty:
            duplicate_indices.update(name_dupes.index)
    
    if duplicate_indices:
        duplicates = df.loc[list(duplicate_indices)]
    
    return duplicates, duplicate_indices


def calculate_data_quality_score(row, email_col='Email', phone_col='Phone', 
                                   name_cols=['First Name', 'Last Name'], company_col='Company',
                                   company_patterns=None, check_pattern_match=False):
    """Calculate a data quality score (0-100) for a row."""
    score = 0
    max_score = 0
    
    # Email (30 points base, +10 bonus for pattern match, -10 penalty for mismatch)
    max_score += 30
    email_bonus = 0
    if email_col in row.index and row[email_col]:
        is_valid, _ = validate_email_format(row[email_col])
        if is_valid:
            score += 30
            
            # Pattern matching bonus/penalty (only if enabled and company has pattern)
            if check_pattern_match and company_patterns:
                company = row.get(company_col, '')
                first_name = row.get('First Name', '')
                last_name = row.get('Last Name', '')
                
                if company and company in company_patterns:
                    matches, detected, expected = check_email_pattern_match(
                        row[email_col], first_name, last_name, company, company_patterns
                    )
                    if matches:
                        email_bonus = 10  # Bonus for matching pattern
                    elif expected:
                        email_bonus = -10  # Penalty for not matching company pattern
    
    # First Name (20 points)
    max_score += 20
    if 'First Name' in row.index and row['First Name']:
        if len(str(row['First Name']).strip()) >= 2:
            score += 20
    
    # Last Name (20 points)
    max_score += 20
    if 'Last Name' in row.index and row['Last Name']:
        if len(str(row['Last Name']).strip()) >= 2:
            score += 20
    
    # Company (15 points)
    max_score += 15
    if company_col in row.index and row[company_col]:
        if len(str(row[company_col]).strip()) >= 2:
            score += 15
    
    # Phone (15 points)
    max_score += 15
    if phone_col in row.index and row[phone_col]:
        phone_clean, is_valid = clean_phone_number(row[phone_col])
        if is_valid:
            score += 15
    
    # Apply email pattern bonus/penalty (can go above 100 or below 0)
    final_score = score + email_bonus
    
    # Normalize to 0-100 range
    normalized_score = int((final_score / max_score) * 100) if max_score > 0 else 0
    normalized_score = max(0, min(100, normalized_score))  # Clamp between 0 and 100
    
    return normalized_score


def clean_data(df, options):
    """Clean data with various options."""
    cleaned_df = df.copy()
    changes = 0
    changed_mask = pd.DataFrame(False, index=df.index, columns=df.columns)
    quality_scores = []
    email_validation_results = []
    pattern_match_info = []
    
    # Detect column names
    email_col = options.get('email_col', 'Email')
    phone_col = options.get('phone_col', 'Phone')
    job_title_col = options.get('job_title_col', 'Job Title')
    company_col = 'Company'
    
    # First pass: clean all data
     for i, row in df.iterrows():
        # Handle NaN values properly - use safe column access
        try:
            first_val = row['First Name'] if 'First Name' in df.columns else ''
        except (KeyError, IndexError):
            first_val = ''
        
        try:
            last_val = row['Last Name'] if 'Last Name' in df.columns else ''
        except (KeyError, IndexError):
            last_val = ''
        
        try:
            company_val = row['Company'] if 'Company' in df.columns else ''
        except (KeyError, IndexError):
            company_val = ''
        
        try:
            email_val = row[email_col] if email_col in df.columns else ''
        except (KeyError, IndexError):
            email_val = ''
        
        orig_first = '' if pd.isna(first_val) else str(first_val).strip()
        orig_last = '' if pd.isna(last_val) else str(last_val).strip()
        orig_company = '' if pd.isna(company_val) else str(company_val).strip()
        email = '' if pd.isna(email_val) else str(email_val).strip()
        
        # First name
        if options.get('clean_names', True):
            first = clean_first_name(orig_first)
        else:
            first = orig_first
        
        # Last name: clean if present, infer from email if missing
        if orig_last:
            if options.get('clean_names', True):
                last = clean_last_name(orig_last)
            else:
                last = orig_last
        else:
            if options.get('infer_last_name', True):
                inferred_last = infer_last_from_email(first, email)
                last = clean_last_name(inferred_last) if inferred_last else ''
            else:
                last = ''
        
        # Company
        if options.get('clean_company', True):
            try:
                company = clean_company(orig_company)
            except Exception as e:
                # If cleaning fails, use original value
                company = orig_company if orig_company else ''
        else:
            company = orig_company
        
        # Email validation
        if email_col in df.columns and options.get('validate_email', True) and email:
            try:
                is_valid, status = validate_email_format(email)
                email_validation_results.append({
                    'index': i,
                    'email': email,
                    'is_valid': is_valid,
                    'status': status
                })
            except Exception as e:
                # If validation fails, mark as invalid
                email_validation_results.append({
                    'index': i,
                    'email': email,
                    'is_valid': False,
                    'status': 'Validation Error'
                })
        
        # Phone cleaning
        if phone_col in df.columns and options.get('clean_phone', True):
            try:
                phone_val = row[phone_col] if phone_col in df.columns else ''
                orig_phone = '' if pd.isna(phone_val) else str(phone_val).strip()
            except (KeyError, IndexError):
                orig_phone = ''
            phone_clean, is_valid = clean_phone_number(orig_phone)
            if phone_clean != str(orig_phone):
                changed_mask.at[i, phone_col] = True
                cleaned_df.at[i, phone_col] = phone_clean
                changes += 1
        
        # Job title cleaning
        if job_title_col in df.columns and options.get('clean_job_title', True):
            try:
                title_val = row[job_title_col] if job_title_col in df.columns else ''
                orig_title = '' if pd.isna(title_val) else str(title_val).strip()
            except (KeyError, IndexError):
                orig_title = ''
            title_clean = clean_job_title(orig_title)
            if title_clean != orig_title:
                changed_mask.at[i, job_title_col] = True
                cleaned_df.at[i, job_title_col] = title_clean
                changes += 1
        
        # Update names and company
        if first != orig_first:
            changed_mask.at[i, 'First Name'] = True
            cleaned_df.at[i, 'First Name'] = first
            changes += 1
        
        if last != orig_last:
            changed_mask.at[i, 'Last Name'] = True
            cleaned_df.at[i, 'Last Name'] = last
            changes += 1
        
        if company != orig_company:
            changed_mask.at[i, 'Company'] = True
            cleaned_df.at[i, 'Company'] = company
            changes += 1
        
    # Analyze company email patterns (if enabled)
    company_patterns = {}
    check_pattern_match = options.get('check_company_email_pattern', False)
    
    if check_pattern_match and 'Company' in cleaned_df.columns:
        company_patterns = analyze_company_email_patterns(
            cleaned_df, email_col, company_col, 'First Name', 'Last Name'
        )
    
    # Second pass: Calculate quality scores with pattern matching
    for i, row in cleaned_df.iterrows():
        if options.get('calculate_quality_score', True):
            quality_score = calculate_data_quality_score(
                row, email_col, phone_col, 
                company_col=company_col,
                company_patterns=company_patterns,
                check_pattern_match=check_pattern_match
            )
            quality_scores.append(quality_score)
            
            # Store pattern match info for reporting
            if check_pattern_match and company_patterns:
                company = row.get(company_col, '')
                email = row.get(email_col, '')
                first_name = row.get('First Name', '')
                last_name = row.get('Last Name', '')
                
                if company and company in company_patterns and email:
                    matches, detected, expected = check_email_pattern_match(
                        email, first_name, last_name, company, company_patterns
                    )
                    if matches is not None:
                        pattern_match_info.append({
                            'index': i,
                            'company': company,
                            'email': email,
                            'matches_pattern': matches,
                            'detected_pattern': detected,
                            'expected_pattern': expected
                        })
    
    # Add quality score column
    if options.get('calculate_quality_score', True) and quality_scores:
        cleaned_df['Quality Score'] = quality_scores
    
    pct = (changes / len(df)) * 100 if len(df) else 0
    return cleaned_df, pct, changed_mask, email_validation_results, pattern_match_info, company_patterns


def split_into_lists_by_company(df, max_lists=4):
    """
    Split contacts into up to max_lists lists so that people
    from the same company are spread across different lists.
    Returns a list of lists of row indices.
    """
    if 'Company' not in df.columns or df.empty:
        return [list(df.index)]

    company_groups = {}
    for idx, company in df['Company'].items():
        company_groups.setdefault(company, []).append(idx)

    max_count = max(len(indices) for indices in company_groups.values())
    num_lists = min(max_lists, max_count if max_count > 0 else 1)

    batches = [[] for _ in range(num_lists)]

    for company, indices in company_groups.items():
        random.shuffle(indices)  # randomise order within company
        for i, idx in enumerate(indices):
            batches[i % num_lists].append(idx)

    return [batch for batch in batches if batch]


def generate_highlighted_excel_with_splits(df, mask, split_batches=None, max_lists=4):
    """
    Create an Excel file with:
      - Sheet "All Contacts": full cleaned data with highlights.
      - If split_batches is provided and not empty:
          up to max_lists sheets "List 1" ... "List N": split by company.
    """
    wb = Workbook()
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

    # ---- Sheet 1: All Contacts ----
    ws_all = wb.active
    ws_all.title = "All Contacts"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:
                col = df.columns[c_idx - 1]
                orig_idx = df.index[r_idx - 2]
                if (col in mask.columns) and mask.at[orig_idx, col]:
                    cell.fill = yellow_fill
                # Highlight low quality scores
                if col == 'Quality Score' and isinstance(value, (int, float)):
                    if value < 50:
                        cell.fill = red_fill

    # ---- Additional sheets: List 1..N (only if splitting enabled) ----
    if split_batches:
        for list_num, indices in enumerate(split_batches[:max_lists], start=1):
            ws = wb.create_sheet(title=f"List {list_num}")
            sub_df = df.loc[indices]
            sub_mask = mask.loc[indices]

            for r_idx, row in enumerate(dataframe_to_rows(sub_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx > 1:
                        col = sub_df.columns[c_idx - 1]
                        orig_idx = sub_df.index[r_idx - 2]
                        if (col in sub_mask.columns) and sub_mask.at[orig_idx, col]:
                            cell.fill = yellow_fill
                        if col == 'Quality Score' and isinstance(value, (int, float)):
                            if value < 50:
                                cell.fill = red_fill

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def detect_columns(df):
    """Auto-detect column names for common variations."""
    column_mapping = {}
    
    # Email detection
    email_patterns = ['email', 'e-mail', 'mail', 'email address']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in email_patterns):
            column_mapping['email_col'] = col
            break
    
    # Phone detection
    phone_patterns = ['phone', 'telephone', 'tel', 'mobile', 'cell']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in phone_patterns):
            column_mapping['phone_col'] = col
            break
    
    # Job title detection
    title_patterns = ['title', 'job title', 'position', 'role']
    for col in df.columns:
        if any(pattern in col.lower() for pattern in title_patterns):
            column_mapping['job_title_col'] = col
            break
    
    return column_mapping


# --- UI Layout ---
st.markdown('<div class="title-text">Cleanr.</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle-text">Clean your data faster.</div>', unsafe_allow_html=True)
st.markdown('<div class="rounded-box">Upload your CSV file and get a cleaned version ready for email outreach.</div>', unsafe_allow_html=True)

# Settings sidebar
with st.sidebar:
    st.header("⚙️ Cleaning Options")
    
    clean_names = st.checkbox("Clean Names", value=True, help="Clean and standardize first and last names")
    clean_company = st.checkbox("Clean Company Names", value=True, help="Clean and standardize company names")
    infer_last_name = st.checkbox("Infer Last Names from Email", value=True, help="Try to infer missing last names from email addresses")
    validate_email = st.checkbox("Validate Emails", value=True, help="Validate email format and flag invalid emails")
    check_company_email_pattern = st.checkbox("Check Company Email Patterns", value=False, 
                                               help="For companies with 2+ contacts, emails matching the company's dominant pattern get higher scores")
    clean_phone = st.checkbox("Clean Phone Numbers", value=True, help="Clean and standardize phone numbers")
    clean_job_title = st.checkbox("Clean Job Titles", value=True, help="Clean and standardize job titles")
    calculate_quality_score = st.checkbox("Calculate Quality Scores", value=True, help="Add a data quality score (0-100) for each row")
    remove_duplicates = st.checkbox("Remove Duplicates", value=False, help="Remove duplicate contacts based on email or name")
    
    st.divider()
    
    split_enabled = st.checkbox(
        "Split by Company",
        value=True,
        help="Split contacts from the same company into separate sending lists"
    )
    
    max_lists = st.slider("Max Lists", 1, 10, 4, help="Maximum number of lists to split contacts into")

uploaded_file = st.file_uploader("Upload CSV File", type=["csv"])

if uploaded_file:
    df = pd.read_csv(uploaded_file, encoding='latin1')
    df.columns = [col.strip().title().replace('_', ' ') for col in df.columns]
    df.rename(columns={'Company Name': 'Company'}, inplace=True)
    
    # Auto-detect columns
    detected_cols = detect_columns(df)
    
    # Prepare options
    options = {
        'clean_names': clean_names,
        'clean_company': clean_company,
        'infer_last_name': infer_last_name,
        'validate_email': validate_email,
        'check_company_email_pattern': check_company_email_pattern,
        'clean_phone': clean_phone,
        'clean_job_title': clean_job_title,
        'calculate_quality_score': calculate_quality_score,
        'email_col': detected_cols.get('email_col', 'Email'),
        'phone_col': detected_cols.get('phone_col', 'Phone'),
        'job_title_col': detected_cols.get('job_title_col', 'Job Title'),
    }
    
    # Clean data
    cleaned_df, percent_cleaned, changed_mask, email_validation, pattern_match_info, company_patterns = clean_data(df, options)
    
    # Handle duplicates
    duplicates_df = pd.DataFrame()
    if remove_duplicates:
        duplicates_df, duplicate_indices = find_duplicates(cleaned_df)
        if not duplicates_df.empty:
            cleaned_df = cleaned_df.drop(duplicates_df.index)
            changed_mask = changed_mask.drop(duplicates_df.index)
            st.warning(f"⚠️ Removed {len(duplicates_df)} duplicate contacts")
    
    # Calculate statistics
    total_rows = len(df)
    cleaned_rows = len(cleaned_df)
    
    # Email statistics
    valid_emails = 0
    invalid_emails = 0
    if email_validation:
        for result in email_validation:
            if result['is_valid']:
                valid_emails += 1
            else:
                invalid_emails += 1
    
    # Quality score statistics
    avg_quality = 0
    if 'Quality Score' in cleaned_df.columns:
        avg_quality = cleaned_df['Quality Score'].mean()
    
    st.success("✅ Done! Your data is cleaned and ready to download.")
    
    # Display statistics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Rows", total_rows)
    with col2:
        st.metric("Rows Cleaned", f"{percent_cleaned:.1f}%")
    with col3:
        if email_validation:
            st.metric("Valid Emails", f"{valid_emails}/{valid_emails + invalid_emails}")
    with col4:
        if avg_quality > 0:
            st.metric("Avg Quality Score", f"{avg_quality:.0f}/100")
    
    # Email validation details
    if email_validation and invalid_emails > 0:
        with st.expander(f"⚠️ {invalid_emails} Invalid Emails Found"):
            invalid_df = pd.DataFrame([
                {'Row': result['index'] + 1, 'Email': result['email'], 'Status': result['status']}
                for result in email_validation if not result['is_valid']
            ])
            st.dataframe(invalid_df, use_container_width=True)
    
    # Duplicates details
    if not duplicates_df.empty and not remove_duplicates:
        with st.expander(f"⚠️ {len(duplicates_df)} Duplicate Contacts Found"):
            st.dataframe(duplicates_df, use_container_width=True)
    
    # Company email pattern analysis
    if check_company_email_pattern and company_patterns:
        matching_count = sum(1 for p in pattern_match_info if p['matches_pattern'])
        non_matching_count = sum(1 for p in pattern_match_info if not p['matches_pattern'])
        
        if pattern_match_info:
            with st.expander(f"📊 Company Email Pattern Analysis ({len(company_patterns)} companies analyzed)"):
                st.info(f"✅ {matching_count} emails match their company's pattern | ⚠️ {non_matching_count} emails don't match")
                
                # Show companies with patterns
                st.markdown("**Companies with detected email patterns:**")
                pattern_summary = []
                for company, info in company_patterns.items():
                    pattern_summary.append({
                        'Company': company,
                        'Pattern': info['pattern'],
                        'Matching Emails': f"{info['count']}/{info['total']}",
                        'Percentage': f"{info['percentage']:.1f}%"
                    })
                st.dataframe(pd.DataFrame(pattern_summary), use_container_width=True)
                
                # Show non-matching emails
                if non_matching_count > 0:
                    st.markdown("**⚠️ Emails that don't match their company's pattern:**")
                    non_matching_df = pd.DataFrame([
                        {
                            'Row': p['index'] + 1,
                            'Company': p['company'],
                            'Email': p['email'],
                            'Detected Pattern': p['detected_pattern'],
                            'Expected Pattern': p['expected_pattern']
                        }
                        for p in pattern_match_info if not p['matches_pattern']
                    ])
                    st.dataframe(non_matching_df, use_container_width=True)
    
    split_batches = None
    
    if split_enabled:
        split_batches = split_into_lists_by_company(cleaned_df, max_lists=max_lists)
        if len(split_batches) > 1:
            st.info(
                f"📧 Splitting is enabled. Your cleaned data has been split into "
                f"{len(split_batches)} sending lists (maximum {max_lists}) to help protect deliverability."
            )
            cols = st.columns(len(split_batches))
            for i, batch in enumerate(split_batches):
                with cols[i]:
                    st.metric(f"List {i+1}", len(batch))
        else:
            st.write("📧 Splitting is enabled, but there are no companies with multiple contacts.")
    
    # Send Usage Log
    cleaned_rows_count = int((percent_cleaned / 100) * len(df))
    usage_data = {
        "type": "usage",
        "sheet": "Usage",
        "filename": uploaded_file.name,
        "rows": len(df),
        "cleaned": cleaned_rows_count,
        "percent_cleaned": round(percent_cleaned, 1),
        "time_saved": round((cleaned_rows_count * 7.5) / 60, 1)
    }
    try:
        requests.post(
            "https://script.google.com/macros/s/AKfycbxM7dmZfMIuWcNWiyxAh8nwX69rvuRaioJ6EH_k7Vx9DRu6DdYdMIO3ZbsZmH--Q5q1/exec",
            json=usage_data,
            timeout=2
        )
    except Exception:
        pass
    
    # Export options
    st.markdown("<div class='section-header'>📥 Download Options</div>", unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Excel export
        excel_file = generate_highlighted_excel_with_splits(
            cleaned_df,
            changed_mask,
            split_batches=split_batches if split_enabled else None,
            max_lists=max_lists
        )
        st.download_button(
            label="📊 Download Excel",
            data=excel_file,
            file_name=uploaded_file.name.replace('.csv', '_cleaned.xlsx'),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    
    with col2:
        # CSV export
        csv_file = cleaned_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📄 Download CSV",
            data=csv_file,
            file_name=uploaded_file.name.replace('.csv', '_cleaned.csv'),
            mime="text/csv",
        )
    
    with col3:
        # JSON export
        json_file = cleaned_df.to_json(orient='records', indent=2).encode('utf-8')
        st.download_button(
            label="📋 Download JSON",
            data=json_file,
            file_name=uploaded_file.name.replace('.csv', '_cleaned.json'),
            mime="application/json",
        )
    
    # Preview
    st.markdown("<div class='section-header'>Preview</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("<b>Before Cleaning</b>", unsafe_allow_html=True)
        st.dataframe(df.head(10), use_container_width=True)
    with col2:
        st.markdown("<b>After Cleaning</b>", unsafe_allow_html=True)
        st.dataframe(cleaned_df.head(10), use_container_width=True)

# --- Feedback Form ---
st.markdown("<div class='feedback-container'>", unsafe_allow_html=True)
st.markdown("<h4>💬 Leave Feedback</h4>", unsafe_allow_html=True)
with st.form(key="feedback_form"):
    feedback = st.text_area("Have any suggestions, bugs, or feature ideas?", height=120, key="feedback_box")
    submitted = st.form_submit_button("Submit")
    if submitted:
        if feedback.strip():
            try:
                requests.post(
                    "https://script.google.com/macros/s/AKfycbxM7dmZfMIuWcNWiyxAh8nwX69rvuRaioJ6EH_k7Vx9DRu6DdYdMIO3ZbsZmH--Q5q1/exec",
                    json={
                        "type": "feedback",
                        "sheet": "Feedback",
                        "timestamp": str(pd.Timestamp.now()),
                        "message": feedback.strip()
                    },
                    timeout=2
                )
                st.success("✅ Thanks! Your feedback was submitted.")
            except Exception as e:
                st.error(f"❌ Failed to submit feedback: {e}")
        else:
            st.warning("✏️ Please write something before submitting.")
st.markdown("</div>", unsafe_allow_html=True)
